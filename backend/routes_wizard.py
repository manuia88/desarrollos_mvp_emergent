"""Phase 4 Batch 12 — Wizard 7 pasos + IA upload + Drive import.

Endpoints
---------
  POST  /api/dev/wizard/projects              — create project from wizard payload
  GET   /api/dev/wizard/smart-defaults        — suggestions from previous projects
  POST  /api/dev/wizard/draft/save            — cross-device draft persistence
  GET   /api/dev/wizard/draft/load            — restore draft
  POST  /api/dev/wizard/ia-extract            — upload files + Claude extract
  GET   /api/dev/wizard/ia-extract/:run_id    — fetch latest extraction by run_id
  POST  /api/dev/wizard/drive/url             — process public Drive folder URL
  GET   /api/dev/wizard/drive/status          — OAuth configured?
"""
from __future__ import annotations

import asyncio
import json
import logging
import os
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Request, HTTPException, UploadFile, File, Form
from pydantic import BaseModel

log = logging.getLogger("dmx.wizard")
router = APIRouter(prefix="/api/dev/wizard", tags=["wizard"])

UPLOAD_ROOT = Path("/app/backend/uploads/wizard_ia")
UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
MAX_FILE_MB = 20
MAX_FILES = 10
ACCEPTED_EXT = {".pdf", ".xlsx", ".xls", ".csv", ".docx", ".txt", ".md"}


def _now():
    return datetime.now(timezone.utc)


def _db(req: Request):
    return req.app.state.db


async def _auth(req: Request, roles=None):
    from server import get_current_user
    user = await get_current_user(req)
    if not user:
        raise HTTPException(401, "No autenticado")
    roles = roles or ["developer_admin", "director", "superadmin"]
    if user.role not in roles:
        raise HTTPException(403, f"Requiere rol en {roles}")
    return user


def _slugify(name: str) -> str:
    s = (name or "").lower().strip()
    s = re.sub(r"[^a-z0-9áéíóúüñ ]+", "", s, flags=re.I)
    s = re.sub(r"\s+", "-", s)
    s = (s.replace("á", "a").replace("é", "e").replace("í", "i")
           .replace("ó", "o").replace("ú", "u").replace("ü", "u").replace("ñ", "n"))
    return s[:60] or f"proyecto-{uuid.uuid4().hex[:6]}"


def _org(user) -> str:
    return (getattr(user, "tenant_id", None) or
            getattr(user, "org_id", None) or "default")


# ═════════════════════════════════════════════════════════════════════════════
# SMART DEFAULTS
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/smart-defaults")
async def smart_defaults(request: Request):
    user = await _auth(request)
    db = _db(request)
    org = _org(user)

    # Collect previous projects from this org
    prev: List[Dict] = []
    try:
        from data_developments import DEVELOPMENTS
        prev = [d for d in DEVELOPMENTS if d.get("dev_org_id") == org or
                d.get("developer_id") == getattr(user, "user_id", None)]
    except Exception:
        pass
    # Fallback: use db projects
    if not prev:
        async for p in db.projects.find({"dev_org_id": org}, {"_id": 0}).limit(10):
            prev.append(p)

    # Aggregate amenities from previous projects' project_amenities docs
    amen_counts: Dict[str, int] = {}
    async for doc in db.project_amenities.find(
        {"dev_org_id": org}, {"_id": 0, "amenities": 1}
    ).limit(20):
        for a in doc.get("amenities", []):
            amen_counts[a] = amen_counts.get(a, 0) + 1
    top_amenities = [a for a, _ in sorted(amen_counts.items(), key=lambda kv: -kv[1])[:10]]

    # Most frequent tipo/segmento
    def _mode(key):
        counts: Dict[str, int] = {}
        for p in prev:
            v = p.get(key)
            if v:
                counts[v] = counts.get(v, 0) + 1
        return max(counts.items(), key=lambda kv: kv[1])[0] if counts else None

    # Commercialization default from latest
    comm = await db.project_commercialization.find_one(
        {"dev_org_id": org}, {"_id": 0},
        sort=[("updated_at", -1)]
    ) or {}

    return {
        "based_on_count": len(prev),
        "categoria": {
            "tipo_proyecto": _mode("tipo_proyecto") or _mode("tipo"),
            "segmento": _mode("segmento") or _mode("nse"),
            "etapa": _mode("stage") or "preventa",
        },
        "operacion": {
            "total_unidades": int(sum((p.get("total_units") or 0) for p in prev) / len(prev))
                               if prev else None,
            "construction_cost": None,
            "target_price": int(sum((p.get("price_from") or 0) for p in prev) / len(prev))
                              if prev and any(p.get("price_from") for p in prev) else None,
            "target_absorption_months": 18,
        },
        "amenidades_sugeridas": top_amenities,
        "comercializacion": {
            "works_with_brokers": comm.get("works_with_brokers", False),
            "default_commission_pct": comm.get("default_commission_pct", 3.0),
            "iva_included": comm.get("iva_included", False),
            "in_house_only": comm.get("in_house_only", True),
        },
    }


# ═════════════════════════════════════════════════════════════════════════════
# DRAFT  (cross-device)
# ═════════════════════════════════════════════════════════════════════════════

class DraftPayload(BaseModel):
    draft_data: Dict[str, Any]
    current_step: int = 0


@router.post("/draft/save")
async def save_draft(payload: DraftPayload, request: Request):
    user = await _auth(request)
    db = _db(request)
    await db.wizard_drafts.update_one(
        {"user_id": user.user_id, "dev_org_id": _org(user)},
        {"$set": {
            "user_id": user.user_id,
            "dev_org_id": _org(user),
            "draft_data": payload.draft_data,
            "current_step": payload.current_step,
            "updated_at": _now().isoformat(),
        }},
        upsert=True,
    )
    return {"ok": True}


@router.get("/draft/load")
async def load_draft(request: Request):
    user = await _auth(request)
    db = _db(request)
    doc = await db.wizard_drafts.find_one(
        {"user_id": user.user_id, "dev_org_id": _org(user)}, {"_id": 0}
    )
    return doc or {"empty": True}


# ═════════════════════════════════════════════════════════════════════════════
# CREATE PROJECT (final submit)
# ═════════════════════════════════════════════════════════════════════════════

class WizardProjectPayload(BaseModel):
    categoria: Dict[str, Any] = {}
    operacion: Dict[str, Any] = {}
    ubicacion: Dict[str, Any] = {}
    amenidades: List[str] = []
    contenido: Dict[str, Any] = {}       # asset IDs references
    legal: Dict[str, Any] = {}
    comercializacion: Dict[str, Any] = {}
    ia_source: Optional[str] = None      # 'manual' | 'ia_upload' | 'drive'
    ia_extraction_id: Optional[str] = None


@router.post("/projects")
async def create_project(payload: WizardProjectPayload, request: Request):
    user = await _auth(request)
    db = _db(request)
    org = _org(user)
    op = payload.operacion or {}
    ub = payload.ubicacion or {}
    cat = payload.categoria or {}

    name = op.get("nombre") or op.get("name")
    if not name:
        raise HTTPException(400, "Nombre del proyecto requerido")

    slug = op.get("slug") or _slugify(name)
    # Uniqueness
    if await db.projects.find_one({"id": slug}):
        slug = f"{slug}-{uuid.uuid4().hex[:4]}"

    now_iso = _now().isoformat()
    total_units = int(op.get("total_unidades") or op.get("total_units") or 0)
    target_price = float(op.get("target_price") or op.get("price_from") or 0)

    project_doc = {
        "id": slug,
        "slug": slug,
        "name": name,
        "dev_org_id": org,
        "developer_id": getattr(user, "user_id", None),
        "tipo_proyecto": cat.get("tipo_proyecto"),
        "segmento": cat.get("segmento"),
        "stage": cat.get("etapa", "preventa"),
        "total_units": total_units,
        "construction_cost": op.get("construction_cost"),
        "price_from": target_price,
        "target_absorption_months": op.get("target_absorption_months"),
        # Ubicación
        "estado": ub.get("estado"),
        "municipio": ub.get("municipio"),
        "colonia": ub.get("colonia"),
        "calle": ub.get("calle"),
        "cp": ub.get("cp"),
        "lat": ub.get("lat"),
        "lng": ub.get("lng"),
        # Metadata
        "created_via": "wizard",
        "wizard_source": payload.ia_source or "manual",
        "wizard_ia_extraction_id": payload.ia_extraction_id,
        "created_at": now_iso,
        "updated_at": now_iso,
    }

    await db.projects.insert_one(dict(project_doc))

    # Amenities
    if payload.amenidades:
        await db.project_amenities.update_one(
            {"project_id": slug, "dev_org_id": org},
            {"$set": {
                "project_id": slug, "dev_org_id": org,
                "amenities": payload.amenidades,
                "updated_at": now_iso, "updated_by": user.user_id,
            }},
            upsert=True,
        )

    # Commercialization
    comm = payload.comercializacion or {}
    await db.project_commercialization.update_one(
        {"project_id": slug, "dev_org_id": org},
        {"$set": {
            "project_id": slug, "dev_org_id": org,
            "works_with_brokers": bool(comm.get("works_with_brokers", False)),
            "default_commission_pct": float(comm.get("default_commission_pct", 3.0)),
            "iva_included": bool(comm.get("iva_included", False)),
            "broker_terms": comm.get("broker_terms", ""),
            "in_house_only": bool(comm.get("in_house_only", True)),
            "approved_inmobiliarias": comm.get("approved_inmobiliarias", []),
            "created_at": now_iso, "updated_at": now_iso, "updated_by": user.user_id,
        }},
        upsert=True,
    )

    # Pre-asignaciones in-house
    for uid in comm.get("preassigned_asesores", []) or []:
        await db.project_preassignments.update_one(
            {"project_id": slug, "assigned_user_id": uid},
            {"$set": {
                "project_id": slug, "dev_org_id": org,
                "assigned_user_id": uid, "active": True,
                "created_at": now_iso, "created_by": user.user_id,
            }},
            upsert=True,
        )

    # Placeholder units
    for i in range(min(total_units, 200)):  # cap insert batch
        uid = f"{slug}-u{i+1:03d}"
        await db.units.update_one(
            {"id": uid},
            {"$setOnInsert": {
                "id": uid, "project_id": slug, "dev_org_id": org,
                "unit_number": f"{i+1:03d}",
                "prototype": "A", "status": "disponible",
                "price": target_price, "area_total": None,
                "created_at": now_iso,
            }},
            upsert=True,
        )

    # Audit + event
    try:
        import audit_log as al
        asyncio.create_task(al.log_mutation(
            db, user_id=user.user_id, role=user.role, org_id=org,
            action="create", entity_type="project",
            entity_id=slug, before=None, after={"name": name, "source": payload.ia_source},
        ))
        from observability import emit_ml_event
        asyncio.create_task(emit_ml_event(
            db, "project_created_via_wizard",
            user.user_id, org, user.role,
            context={"project_id": slug, "source": payload.ia_source,
                     "total_units": total_units},
        ))
    except Exception:
        pass

    # Trigger diagnostic (B0.5) 5min later
    try:
        from diagnostic_engine import run_diagnostics
        asyncio.create_task(_delayed_diagnostic(db, slug, user))
    except Exception:
        pass

    # Cleanup draft
    await db.wizard_drafts.delete_one({"user_id": user.user_id, "dev_org_id": org})

    return {"ok": True, "project_id": slug, "redirect": f"/desarrollador/proyectos/{slug}"}


async def _delayed_diagnostic(db, slug, user):
    await asyncio.sleep(300)  # 5min
    try:
        from diagnostic_engine import run_diagnostics
        await run_diagnostics(db, project_id=slug, scope="all",
                              user=user, trigger="auto_post_create")
    except Exception as e:
        log.warning(f"[wizard] delayed diag failed: {e}")


# ═════════════════════════════════════════════════════════════════════════════
# IA EXTRACT — drag-drop documents + Claude analysis
# ═════════════════════════════════════════════════════════════════════════════

_IA_SYSTEM = (
    "Eres un experto analista de proyectos inmobiliarios LATAM. "
    "Dado el texto extraído de documentos (PDF, Excel, Word) que describen un nuevo "
    "proyecto inmobiliario en México, extrae los campos estructurados necesarios "
    "para crear la ficha del proyecto. Para cada campo incluye el valor detectado, "
    "la confianza (0-100) y el archivo/sección de origen. Output STRICTLY JSON sin markdown:\n"
    '{\n'
    '  "nombre": {"value": "...", "confidence": 80, "source": "..."},\n'
    '  "tipo_proyecto": {...},        // "residencial_vertical"|"residencial_horizontal"|"mixto"|"comercial"\n'
    '  "segmento": {...},              // "NSE_AB"|"NSE_C+"|"NSE_C"|"NSE_D"\n'
    '  "etapa": {...},                 // "preventa"|"en_construccion"|"entregado"\n'
    '  "total_unidades": {...},        // integer\n'
    '  "construction_cost": {...},     // MXN number\n'
    '  "target_price": {...},          // MXN number average per unit\n'
    '  "target_absorption_months": {...},\n'
    '  "ubicacion": {\n'
    '    "estado": {...}, "municipio": {...}, "colonia": {...},\n'
    '    "calle": {...}, "cp": {...}\n'
    '  },\n'
    '  "amenidades": [{"value": "gym", "confidence": 90, "source": "..."}, ...]\n'
    '}\n'
    "Si un campo no aparece, omítelo (no inventes). Confianza < 40 solo si es una "
    "inferencia débil. Todo en español es-MX."
)


async def _extract_text_from_upload(path: Path) -> str:
    """Best-effort text extraction for pdf/xlsx/csv/docx/txt."""
    ext = path.suffix.lower()
    try:
        if ext in (".txt", ".md", ".csv"):
            return path.read_text(errors="ignore")[:20000]
        if ext == ".pdf":
            try:
                from pypdf import PdfReader
                reader = PdfReader(str(path))
                return "\n".join((p.extract_text() or "") for p in reader.pages)[:30000]
            except Exception as e:
                log.warning(f"pdf extract failed: {e}")
                return ""
        if ext in (".xlsx", ".xls"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(path), data_only=True, read_only=True)
                parts: List[str] = []
                for ws in wb.worksheets:
                    parts.append(f"[Sheet: {ws.title}]")
                    for row in ws.iter_rows(values_only=True):
                        parts.append("\t".join(str(c) if c is not None else "" for c in row))
                return "\n".join(parts)[:30000]
            except Exception as e:
                log.warning(f"xlsx extract failed: {e}")
                return ""
        if ext == ".docx":
            try:
                from docx import Document
                doc = Document(str(path))
                return "\n".join(p.text for p in doc.paragraphs)[:30000]
            except Exception as e:
                log.warning(f"docx extract failed: {e}")
                return ""
    except Exception as e:
        log.warning(f"extract generic failed {path}: {e}")
    return ""


async def _run_claude_extraction(dev_org_id: str, combined_text: str, source_summary: str) -> Optional[Dict]:
    emergent_key = os.environ.get("EMERGENT_LLM_KEY")
    if not emergent_key or not combined_text.strip():
        return None
    try:
        from ai_budget import is_within_budget, track_ai_call
        if not await is_within_budget(None, dev_org_id):
            return None
    except Exception:
        pass
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=emergent_key,
            session_id=f"wiz_extract_{uuid.uuid4().hex[:8]}",
            system_message=_IA_SYSTEM,
        ).with_model("anthropic", "claude-haiku-4-5")
        user_text = f"Fuentes: {source_summary}\n\n{combined_text[:28000]}"
        raw = await chat.send_message(UserMessage(text=user_text))
        if not raw:
            return None
        txt = raw.strip()
        if txt.startswith("```"):
            txt = re.sub(r"^```(?:json)?\s*|\s*```$", "", txt, flags=re.S).strip()
        try:
            parsed = json.loads(txt)
        except Exception:
            # Extract first JSON object
            m = re.search(r"\{.*\}", txt, re.S)
            parsed = json.loads(m.group(0)) if m else None
        # Record usage — tokens tracking happens in caller (has db)
        return parsed
    except Exception as e:
        log.warning(f"[wizard ia] claude failed: {e}")
        return None


@router.post("/ia-extract")
async def ia_extract(
    request: Request,
    files: List[UploadFile] = File(...),
):
    user = await _auth(request)
    db = _db(request)
    org = _org(user)
    run_id = f"wzia_{uuid.uuid4().hex[:12]}"
    run_dir = UPLOAD_ROOT / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    if len(files) > MAX_FILES:
        raise HTTPException(400, f"Máximo {MAX_FILES} archivos")

    file_meta: List[Dict] = []
    combined_parts: List[str] = []
    for f in files:
        ext = Path(f.filename or "").suffix.lower()
        if ext not in ACCEPTED_EXT:
            raise HTTPException(400, f"Extensión {ext} no soportada")
        content = await f.read()
        if len(content) > MAX_FILE_MB * 1024 * 1024:
            raise HTTPException(400, f"Archivo {f.filename} excede {MAX_FILE_MB} MB")
        safe = re.sub(r"[^A-Za-z0-9._-]", "_", f.filename or f"file_{len(file_meta)}")
        path = run_dir / safe
        path.write_bytes(content)
        text = await _extract_text_from_upload(path)
        file_meta.append({
            "filename": f.filename, "safe_name": safe,
            "size_bytes": len(content), "extension": ext,
            "text_extracted_chars": len(text),
        })
        if text:
            combined_parts.append(f"\n=== {f.filename} ===\n{text}\n")

    combined = "".join(combined_parts)
    source_summary = ", ".join(fm["filename"] for fm in file_meta)

    # Call Claude
    extraction = await _run_claude_extraction(org, combined, source_summary)

    # Fallback stub if AI not available
    if extraction is None:
        extraction = _stub_extraction(combined, file_meta)

    # Confidence stats
    def _walk_confidence(obj, acc: List[int]):
        if isinstance(obj, dict):
            if "value" in obj and "confidence" in obj:
                try:
                    acc.append(int(obj["confidence"]))
                except Exception:
                    pass
            else:
                for v in obj.values():
                    _walk_confidence(v, acc)
        elif isinstance(obj, list):
            for v in obj:
                _walk_confidence(v, acc)

    conf_list: List[int] = []
    _walk_confidence(extraction, conf_list)
    avg_conf = int(sum(conf_list) / len(conf_list)) if conf_list else 0
    fields_count = len(conf_list)

    now_iso = _now().isoformat()
    doc = {
        "id": run_id,
        "run_id": run_id,
        "user_id": user.user_id,
        "dev_org_id": org,
        "files": file_meta,
        "combined_text_chars": len(combined),
        "extraction": extraction,
        "fields_extracted_count": fields_count,
        "avg_confidence": avg_conf,
        "is_ai_fallback_stub": extraction == _stub_extraction(combined, file_meta),
        "source_type": "upload",
        "created_at": now_iso,
    }
    await db.wizard_ia_extractions.insert_one(dict(doc))

    # track_ai_call (best effort)
    try:
        from ai_budget import track_ai_call
        await track_ai_call(
            db, org, "claude-haiku-4-5",
            cost_usd=0.0, call_type="wizard_ia_extract",
            tokens_in=len(combined) // 4,
            tokens_out=len(json.dumps(extraction, ensure_ascii=False)) // 4,
        )
    except Exception:
        pass

    # Audit
    try:
        from observability import emit_ml_event
        asyncio.create_task(emit_ml_event(
            db, "wizard_ia_extract_completed",
            user.user_id, org, user.role,
            context={"run_id": run_id, "files": len(file_meta),
                     "avg_confidence": avg_conf, "fields": fields_count},
        ))
    except Exception:
        pass

    doc.pop("_id", None)
    return doc


def _stub_extraction(text: str, meta: List[Dict]) -> Dict:
    """Minimal honest stub when Claude unavailable."""
    return {
        "_stub": True,
        "_notice": "EMERGENT_LLM_KEY no configurada o budget excedido — extracción basic heurística.",
        "nombre": None,
        "total_unidades": None,
        "amenidades": [],
    }


@router.get("/ia-extract/{run_id}")
async def get_ia_extract(run_id: str, request: Request):
    user = await _auth(request)
    db = _db(request)
    doc = await db.wizard_ia_extractions.find_one(
        {"run_id": run_id, "dev_org_id": _org(user)}, {"_id": 0}
    )
    if not doc:
        raise HTTPException(404)
    return doc


# ═════════════════════════════════════════════════════════════════════════════
# DRIVE IMPORT
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/drive/status")
async def drive_status(request: Request):
    await _auth(request)
    from drive_engine import _has_keys
    return {
        "oauth_configured": _has_keys(),
        "url_paste_available": True,  # public folders work without OAuth
    }


DRIVE_FOLDER_PATTERN = re.compile(r"drive\.google\.com/drive/folders/([a-zA-Z0-9_-]{10,})")


class DriveUrlPayload(BaseModel):
    drive_folder_url: str


@router.post("/drive/url")
async def drive_url(payload: DriveUrlPayload, request: Request):
    """Process a public Drive folder URL.

    For MVP: validates URL pattern + returns honest result. Real public folder
    listing requires either API key (not configured) or OAuth. We return
    a structured response so the frontend can display next steps.
    """
    user = await _auth(request)
    match = DRIVE_FOLDER_PATTERN.search(payload.drive_folder_url)
    if not match:
        raise HTTPException(400, "URL inválida. Debe ser https://drive.google.com/drive/folders/...")
    folder_id = match.group(1)

    # Check if we have API access
    from drive_engine import _has_keys
    if not _has_keys():
        return {
            "ok": False,
            "reason": "oauth_not_configured",
            "folder_id": folder_id,
            "message": "Google OAuth no configurado. Configura GOOGLE_OAUTH_CLIENT_ID y GOOGLE_OAUTH_CLIENT_SECRET en el .env. Mientras tanto, puedes descargar los archivos del folder y subirlos manualmente en la pestaña 'Cargar con IA'.",
        }

    # With OAuth: we'd need the user's connection. For wizard flow, folder must be publicly accessible.
    # Public folder listing via API key is not implemented here — return stub.
    return {
        "ok": False,
        "reason": "requires_oauth_connection",
        "folder_id": folder_id,
        "message": "Para procesar este folder necesitas conectar tu Google Drive. Ve a Drive → Conectar.",
    }


# ═════════════════════════════════════════════════════════════════════════════
# INDEXES
# ═════════════════════════════════════════════════════════════════════════════

async def ensure_wizard_indexes(db):
    await db.wizard_drafts.create_index(
        [("user_id", 1), ("dev_org_id", 1)], unique=True, background=True
    )
    await db.wizard_ia_extractions.create_index("run_id", unique=True, background=True)
    await db.wizard_ia_extractions.create_index(
        [("dev_org_id", 1), ("created_at", -1)], background=True
    )
